"""
    位号合并
"""

from openpyxl import load_workbook

import time

# 打开目标工作簿
work_book = load_workbook('./location_merge.xlsx')
# 获取存放数据的工作表
work_sheet_source = work_book['Sheet1']
# 获取目标写入数据的工作表
work_sheet_result = work_book['data']
# 清空要写入数据的工作表数据
work_sheet_result.delete_rows(1, work_sheet_result.max_row)
# 获取待处理数据内容
data_list = list(work_sheet_source.iter_rows(min_row=1, min_col=1, values_only=True))  # type: list
# 定义字典, 以key为apn, value为位号
apn_dict = {}  # type: dict[str: str]
# 定义字典, 以key为apn, value为用量
num_dict = {}  # type: dict[str: int]
for data in data_list:
    # 判断改apn在字典中是否存在
    if data[1] in apn_dict.keys():
        num_dict[data[1]] = num_dict[data[1]] + 1
        apn_dict[data[1]] = apn_dict[data[1]] + ",{}".format(data[5])
    else:
        # apn在字典中不存在
        num_dict[data[1]] = 1
        apn_dict[data[1]] = data[5]
print(num_dict)
print(apn_dict)
# 定义列表, 存放转换为列表的源数据
list_switch = []  # type: list
for data in data_list:
    list_switch.append(list(data))
# 遍历列表源数据, 修改位号和用量
for index in range(len(data_list)):
    if index != 0:
        # 修改用量
        list_switch[index][4] = num_dict[list_switch[index][1]]
        # 修改位号
        list_switch[index][5] = apn_dict[list_switch[index][1]]
        print(list_switch[index])
# 创建字典, 去重
dict_result = {}
for result in list_switch:
    dict_result[result[1]] = result
for result in dict_result:
    print(dict_result[result])
    work_sheet_result.append(dict_result[result])
# 保存工作簿
work_book.save('./location_merge.xlsx')
print('数据处理完毕')
time.sleep(3)
