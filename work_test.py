import re
import time

from openpyxl import load_workbook

if __name__ == '__main__':
    # 打开目标工作簿
    workbook = load_workbook(input('请输入目标工作簿路径:'))
    # 获取目标工作表
    worksheet = workbook[input('请输入数据源工作表名:')]
    # 新建工作表存储数据
    new_work_sheet = workbook.create_sheet(input('请输入要存储处理后数据工作表名:'))
    # 新建工作表添加表头
    new_work_sheet.append(('Bose-PN', 'Vendor', 'Bose-MPN'))
    # 获取数据
    data_list = worksheet.iter_rows(min_row=2, values_only=True)
    for data in data_list:
        # print(data)
        # 拆分厂商数据为列表
        vendor_data_list = data[1].split('\n')
        # 循环遍历每一个厂商的数据
        for vendor_data in vendor_data_list:
            # 去除数据中的 EOL 后缀
            vendor_data = vendor_data.replace('(EOL)', '')
            try:
                # 定义变量存储数据中厂商和MPN的个数
                num = vendor_data.count('--')
                if num == 1:
                    # 分别提取厂商和MPN信息
                    vendor = vendor_data.split('--')[0].rstrip().lstrip()
                    MPN = vendor_data.split('--')[1].strip()
                    if 'mm' in MPN:
                        # 使用正则表达式匹配并去除MPN中的无用信息
                        MPN = re.compile('\(.*\)').sub('', MPN).strip()
                    print('{} ==> {}:{}'.format(data[0], vendor, MPN))
                    # 添加数据到工作表
                    if vendor_data == vendor_data_list[0]:
                        new_work_sheet.append((data[0], vendor, MPN))
                    else:
                        new_work_sheet.append(('', vendor, MPN))
                else:
                    # 该数据中不止一个厂商, 需要特别处理
                    vendor1 = vendor_data.split('--')[0].rstrip().lstrip()
                    MPN1 = vendor_data.split('--')[1].split('(')[0].strip()
                    if 'mm' in MPN1:
                        MPN1 = re.compile('\(.*\)').sub('', MPN1).strip()
                    vendor2 = vendor_data.split('--')[1].split('(')[1]
                    MPN2 = vendor_data.split('--')[2].replace(')', '')
                    if 'mm' in MPN2:
                        MPN = re.compile('\(.*\)').sub('', MPN2).strip()
                    print('{} ==> {}:{}, {}:{}'.format(data[0], vendor1, MPN1, vendor2, MPN2))
                    # 添加数据到工作表
                    if vendor_data == vendor_data_list[0]:
                        new_work_sheet.append((data[0], vendor1, MPN1, vendor2, MPN2))
                    else:
                        new_work_sheet.append(('', vendor1, MPN1, vendor2, MPN2))
            except:
                pass
    # 保存工作簿
    workbook.save('./operatior_data/10-22/881046-0010 SERENA EARBUDS,BLACK,WW BOM Report.xlsx')
    print('数据处理完毕')
    time.sleep(3)
