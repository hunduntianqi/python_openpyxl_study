"""
    openpyxl操作单元格:
        工作表中的每一个格子都是一个单元格对象
        获取单元格对象:
            1. 工作表对象的 iter_rows() 方法:
                a. 先通过iter_rows()方法获取存储单元格对象的可迭代对象
                b. 遍历可迭代对象获取每一行数据的元组
                c. 遍历元组获取每一个单元格对象
            2. 先获取整行或整列数据的元组, 再遍历元组获取单元格对象
                a. 获取包含 整行 / 整列 数据的元组
                b. 遍历元组获取单元格对象
            3. 通过单元格坐标获取单元格对象
                cell_object = worksheet_object['cell_coord']
        单元格对象的基本操作:
            1. value属性:
                获取单元格的数据值 ==> cell_value = cell_object.value
                给单元格对象赋值 ==> cell_object.value = value
"""
# 导入模块
from openpyxl import load_workbook

if __name__ == '__main__':
    # 打开工作簿
    workbook = load_workbook('../operatior_data/codes/material/practice2_copy.xlsx')
    # 获取 '下半年公司名单' 工作表
    worksheet = workbook['下半年公司名单']
    # 获取 '部门' 一列数据, 该列为 'D' 列
    data = worksheet['D']
    # 遍历数据获取单元格对象
    for cell in data:
        # 判断是否为表头部分
        if cell.value == '部门':
            continue
        else:
            # 将部门统一修改为 战略储备部
            cell.value = '战略储备部'
    # 保存工作簿
    workbook.save('../operatior_data/codes/material/practice2_copy.xlsx')
