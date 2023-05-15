"""
    数据匹配:
        在多个表之间匹配相关联数据
        使用场景:
            要处理的工作表中存在相关联数据, 根据关联将两张表中的数据链接起来
            比如excel中的 vlookup() 函数
        数据匹配步骤:
            1. 获取数据:
                根据需求确定数据范围, 以此来决定获取哪些单元格对象处理数据
            2. 使用数据:
                a. 明确多张表格之间的关联关系, 找到相关联的数据
                b. 将数据按照需要存储起来(推荐使用字典, 关联数据作为值), 然后使用数据与其他表格数据进行匹配
            3. 数据输出:
                根据需求将数据处理结果进行处理, 保存到文件中或打印控制台显示
    数据匹配案例:
        1. 获取到十月考勤统计表中的员工迟到次数信息
        2. 与迟到次数月度统计(10月更新)工作薄中的员工迟到信息匹配核对
"""
# 导入模块
from openpyxl import load_workbook

if __name__ == '__main__':
    # 打开十月考勤信息统计表
    oct_workbook = load_workbook('../operatior_data/codes/material/10月考勤统计.xlsx')
    # 获取活动工作表
    oct_worksheet = oct_workbook.active
    # 创建字典添加员工迟到次数信息
    info_dict = {}
    # 循环遍历给字典添加数据
    for data in oct_worksheet.iter_rows(min_row=2, values_only=True):
        # 获取员工工号
        info_num = data[0]
        # 获取员工迟到次数
        delay_num = data[4]
        # 将数据添加到字典
        info_dict[info_num] = delay_num
    print(info_dict)
    # 打开迟到次数月度统计(10月更新)工作簿
    delay_workbook = load_workbook('../operatior_data/codes/material/迟到次数月度统计（10月更新）.xlsx')
    # 获取活动工作表
    delay_worksheet = delay_workbook.active
    # 循环比对迟到次数数据
    for delay_row in delay_worksheet.iter_rows(min_row=3, max_col=13, values_only=True):
        # 获取该表中员工十月迟到次数
        delay_num = delay_row[12]
        # 以工号为键, 从字典中取值比对两表中迟到次数是否一致
        if info_dict[delay_row[0]] != delay_num:
            print('员工({}){}迟到次数信息不一致, 请核对！！'.format(delay_row[0], delay_row[1]))
