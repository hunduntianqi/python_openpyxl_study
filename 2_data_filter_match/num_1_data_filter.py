"""
    数据筛选:
        指在一定量的数据中根据需要筛选出自己需要的数据
        筛选步骤:
            1. 获取数据:
                根据需求确定数据范围, 以此来决定获取哪些单元格对象处理数据
            2. 筛选数据:
                明确筛选条件, 使用相关知识筛选数据 (条件判断语句, 比较运算符, 成员运算符和逻辑运算符等)
            3. 数据输出:
                根据需求将数据处理结果进行处理, 保存到文件中或打印控制台显示
    数据筛选案例:
        筛选迟到人员信息:
            在【10月考勤统计.xlsx】工作簿中, 保存了公司一百名员工的迟到信息, 这些信息包含了迟到时间和迟到次数;
            公司规定, 迟到时间超过45分钟且迟到过3次以上的员工记为考勤不合格, 需要扣除300的考勤保证金
"""
# 导入模块
from openpyxl import load_workbook

if __name__ == '__main__':
    # 打开 '十月考勤统计.xlsx' 工作簿
    workbook = load_workbook('../operatior_data/codes/material/10月考勤统计.xlsx')
    # 获取活动工作表
    work_sheet = workbook.active
    # 定义列表存储考勤不合格人员信息
    list_data: list[tuple] = []
    # 获取表头数据
    header_data = work_sheet[1]
    # 定义列表存储表头数据值
    list_header: list = []
    # 遍历表头数据
    for data in header_data:
        list_header.append(data.value)
    # 表头数据添加到列表中
    list_data.append(tuple(list_header))
    # 获取除表头数据外的所有数据
    iter_data = work_sheet.iter_rows(min_row=2, values_only=True)
    # 遍历数据可迭代对象, 将元祖数据添加到列表中
    for staff_data in iter_data:
        # 判断考勤是否合格
        if staff_data[3] > 45 and staff_data[4] > 3:
            # 将不合格人员信息存储到列表中
            list_data.append(staff_data)
    # 打印考勤不合格人员信息
    for delay_data in list_data:
        print('{}\t{}\t{}\t\t\t{}\t\t\t{}'.format(
            delay_data[0], delay_data[1], delay_data[2], delay_data[3], delay_data[4]))
