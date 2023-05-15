import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border

if __name__ == '__main__':
    # 定义表头颜色样式为橙色, 创建颜色填充对象
    header_fill = PatternFill('solid', fgColor='FF7F24')
    # 定义表中颜色样式为淡黄色, 创建颜色填充对象
    content_fill = PatternFill('solid', fgColor='FFFFE0')
    # 定义表尾颜色样式为淡桔红色, 创建颜色填充对象
    bottom_fill = PatternFill('solid', fgColor='EE9572')

    # 定义对齐样式横向居中、纵向居中
    align = Alignment(horizontal='center', vertical='center')

    # 定义单元格线条样式为细条
    side = Side('thin')
    # 定义表头边框样式，有底边和右边
    header_border = Border(bottom=side, right=side)
    # 定义表中、表尾边框样式，有左边和右边
    content_border = Border(left=side, right=side)
    # 定义表尾边框样式，有底边和右边
    bottom_border = Border(bottom=side, right=side)
    # 定义源文件夹路径
    path_source = '../operatior_data/codes/material/各部门利润表汇总/'
    # 定义副本文件存储文件夹路径
    path_result = '../operatior_data/codes/material/各部门利润表汇总-副本/'
    # 创建存储副本文件的文件目录
    try:
        os.mkdir(path_result)
    except:
        print("目标文件夹已存在！！")
    # 获取源文件夹下所有文件名列表
    list_file_name = os.listdir(path_source)
    # 循环遍历获取源文件名
    for file_name in list_file_name:
        # 拼接源文件路径
        file_source_path = path_source + file_name
        # 打开源文件工作簿
        source_workbook = load_workbook(file_source_path)
        # 获取活动工作表
        source_worksheet = source_workbook.active
        # 调整列宽
        source_worksheet.column_dimensions['A'].width = 15
        source_worksheet.column_dimensions['B'].width = 25
        source_worksheet.column_dimensions['C'].width = 50
        source_worksheet.column_dimensions['D'].width = 10
        source_worksheet.column_dimensions['E'].width = 20
        source_worksheet.column_dimensions['F'].width = 15
        # 调整表头样式
        for cell in source_worksheet[1]:
            # 设置单元格填充颜色
            cell.fill = header_fill
            # 设置单元格对齐方式
            cell.alignment = align
            # 设置单元格边框
            cell.border = header_border
        # 获取最后一行行号
        max_row_num = source_worksheet.max_row
        # 调整表中样式
        for row in source_worksheet.iter_rows(min_row=2, max_row=(max_row_num - 1)):
            # 循环取出单元格，调整表中样式
            for cell in row:
                cell.fill = content_fill
                cell.alignment = align
                cell.border = content_border
        # 调整表尾样式
        for cell in source_worksheet[max_row_num]:
            cell.fill = bottom_fill
            cell.alignment = align
            cell.border = bottom_border
        # 保存副本文件
        source_workbook.save(path_result + file_name.split('.')[0] + '-副本.xlsx')
