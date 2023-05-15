"""
    openpyxl操作工作簿:
        一个 '.xlsx' 格式的Excel文件就代表了一个工作簿对象
        获取工作簿对象:
            from openpyxl import load_workbook, Workbook
            打开已存在工作簿:
                workbook_object = load_workbook(file_path)
            创建新的工作簿对象:
                workbook_object = Workbook() ==> 可以创建一个空的工作簿
            保存工作簿到本地 ==> workbook_object.save(file_path)
        获取工作簿中的活动工作表对象:
            worksheet_object = workbook_object.active ==> 返回工作簿中的活动工作表
        获取工作簿中的指定工作表:
            worksheet_object = workbook_object['sheet_name']
        创建新的工作表:
            workbook_object.create_sheet('new_sheet_name')
        删除工作表:
            删除工作表操作, 要先获取到工作表对象, 才可以删除工作表
            1. 获取工作表对象:
                worksheet_object = workbook_object['sheet_name']
            2. 删除工作表:
                workbook_object.remove(worksheet_object)
        复制工作表:
            复制工作表操作, 要先获取到工作表对象, 才可以复制工作表
            1. 获取工作表对象:
                worksheet_object = workbook_object['sheet_name']
            2. 复制工作表:
                workbook_object.copy_worksheet(worksheet_object)
"""
# 导入模块
from openpyxl import load_workbook, Workbook

if __name__ == '__main__':
    # 打开已存在的工作簿
    workbook_old = load_workbook('../operatior_data/codes/practice1.xlsx')
    # 打印工作簿对象
    print(workbook_old)
    # 获取活动工作表
    work_sheet = workbook_old.active
    print(work_sheet)
    # 创建新的工作簿
    workbook_new = Workbook()
    # 打印新的工作簿
    print(workbook_new)
    # 保存新建的工作簿
    workbook_new.save('../operatior_data/codes/practice1_new.xlsx')
