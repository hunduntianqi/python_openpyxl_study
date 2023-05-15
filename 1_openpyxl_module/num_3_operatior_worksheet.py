"""
    openpyxl操作工作表:
        excel工作簿中每一页都是一个工作表
        获取工作表对象:
            获取活动工作表 ==> worksheet_object = workbook_object.active
            根据表名获取指定工作表 ==> worksheet_object = workbook_object['sheet_name
        修改工作表名称 ==> worksheet_object.title = 'new_sheet_name'
        工作表中冻结窗口 ==> worksheet_object.freeze_panes = 'cell_name', 例: sheet.freeze_panes = 'C3'
        添加筛选器:
            worksheet_object.auto_filter.ref = worksheet_object.dimension ==> 给所有字段添加筛选器
            worksheet_object.auto_filter.ref = 'cell_coord' ==> 指定单元格坐标添加筛选器
        工作表对象基本操作:
            获取 单行 / 单列 数据:
                根据行数获取单行数据 ==> tuple_object = worksheet_object[line_num]
                根据列名获取单列数据 ==> tuple_object = worksheet_object['column_name']
                获取最后一行非空数据 ==> tuple_object = worksheet_object[worksheet_object.max_row]
                    worksheet_object.max_row ==> 可以获取工作表中的最后一行非空数据行数
                以上方法会以元组形式返回包含 列 或 行 所有非空单元格的数据
            获取多行数据 ==> iter_rows()方法:
                语法: worksheet_object.iter_rows(min_row, max_row, min_col, max_col, values_only = False)
                    参数说明:
                        min_row / max_row: 最小行索引 / 最大行索引
                        min_col / max_col: 最小列索引 / 最大列索引
                            最小列索引默认值为 1
                            最大列索引默认值为 '非空单元格' 的最后一列
                        values_only: 决定是否返回单元格的值, 默认值为 False 表示不返回值, 返回单元格对象
                                     只读数据时可以将其设置为True, 使其返回单元格的具体值
                    方法返回值:
                        为一个可迭代对象, 包含有 N 个元组, N为指定范围的行数, 每个元组中都存储有一行数据
                        for循环遍历可以取出每一行数据, 如果指定的行中没有数据, 就会返回一个空的元组(None, None)
            添加数据 ==> append():
                语法:
                    worksheet_object.append(list / tuple)
                注意: 对工作簿中数据进行更改后, 要记得保存工作簿
            插入空行和空列:
                插入空行 ==> worksheet_object.insert_rows(line_index, amount=line_num)
                    在行数 'line_index' 下方插入 line_num 行的空行
                插入空列 ==> worksheet_object.insert_cols(column_index, amount=column_num)
                    在列数 'column_index' 左侧插入 line_num 行的空行
            删除行和列:
                删除行 ==> worksheet_object.delete_rows(line_index, amount=line_num)
                    从行数 'line_index' 开始删除指定的行数
                删除列 ==> worksheet_object.delete_cols(column_index, amount=column_num)
                    从列数 'column_index' 开始, 删除指定的列数
            移动单元格数据:
                worksheet_object.move_range('data_area', rows=line_num, cols=column_num)
                    参数说明:
                        data_area: 指定要移动数据的范围, 例: 'D2:G5'
                        rows: 指定要移动数据的行数, 正数为向下移动, 负数为向上移动
                        cols: 指定要移动数据的列数, 正数为向右移动, 负数为向左移动
"""
# 导入openpyxl相关模块
from openpyxl import load_workbook

if __name__ == '__main__':
    # 打开工作表
    workbook = load_workbook('../operatior_data/codes/practice1.xlsx')
    # 获取活动工作表对象
    worksheet = workbook.active
    # 获取工作表前 5 行数据
    iter_data_list = worksheet.iter_rows(min_row=1, max_row=5, values_only=True)
    print(iter_data_list)
    # 读取数据
    for iter_data in iter_data_list:
        print(iter_data)
    # 添加数据
    worksheet.append(['S1911', '萧爵瑟', 3000, '内容'])
    # 保存工作簿
    workbook.save('../operatior_data/codes/practice1.xlsx')
