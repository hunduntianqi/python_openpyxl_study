
"""
    位号拆分
"""

from openpyxl import load_workbook

import time

# 打开目标工作簿
work_book = load_workbook('./location_split.xlsx')
# 获取存放数据的工作表
work_sheet_source = work_book['Sheet1']
# 获取目标写入数据的工作表
work_sheet_result = work_book['data']
# 清空要写入数据的工作表数据
work_sheet_result.delete_rows(1, work_sheet_result.max_row)
# 获取待处理数据内容
data_list = work_sheet_source.iter_rows(min_row=1, min_col=1, values_only=True)
for data in data_list:
    print(data)
    # 判断是否为表头
    if "Location" in data:
        # 先写入表头
        work_sheet_result.append(data)
    else:
        # 获取位号列数据, 判断是否需拆分位号
        if "," in data[5]:
            # 拆分位号, 返回列表
            location_list = data[5].split(",")
            # 遍历列表, 写入数据
            for location in location_list:
                # 整理数据
                location_tulpe = (data[0], data[1], data[2], data[3], 1, location, data[6], data[7],
                                  data[8], data[9], data[10])
                work_sheet_result.append(location_tulpe)
        else:
            # 不需要拆分位号, 直接写入数据
            work_sheet_result.append(data)
# 保存工作簿
work_book.save("./location_split.xlsx")
print('数据处理完毕')
time.sleep(3)
